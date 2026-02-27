from __future__ import annotations

import os
from dataclasses import dataclass
from functools import lru_cache
from typing import Optional, Dict, Tuple

from openpyxl import load_workbook
from rapidfuzz import fuzz, process


def _norm(s: str) -> str:
    return " ".join((s or "").lower().strip().split())


@dataclass
class PriceHit:
    name: str
    price_raw: Optional[float]
    score: int
    source: str


class PriceIndex:
    def __init__(self, xlsx_dir: str):
        self.xlsx_dir = xlsx_dir
        self._loaded = False
        self._names: list[str] = []
        self._prices: Dict[str, float] = {}
        self._source_by_name: Dict[str, str] = {}

    def load(self):
        if self._loaded:
            return
        self._loaded = True

        if not os.path.isdir(self.xlsx_dir):
            return

        for fn in os.listdir(self.xlsx_dir):
            if not fn.lower().endswith(".xlsx"):
                continue
            path = os.path.join(self.xlsx_dir, fn)
            try:
                self._ingest_xlsx(path, fn)
            except Exception:
                continue

        self._names = list(self._prices.keys())

    def _ingest_xlsx(self, path: str, source_name: str):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        # Find header row
        headers = None
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if not row:
                continue
            row_s = [str(c).strip().lower() if c is not None else "" for c in row]
            if any(k in row_s for k in ["card", "name", "product", "title"]):
                headers = row_s
                break

        if not headers:
            return

        def col_idx(keys: list[str]) -> Optional[int]:
            for k in keys:
                if k in headers:
                    return headers.index(k)
            return None

        name_i = col_idx(["card", "name", "product", "title"]) or 0

        # Price columns vary wildly; try common choices
        price_i = None
        for key in ["price", "raw", "loose price", "loose", "ungraded", "ungraded price", "market price"]:
            if key in headers:
                price_i = headers.index(key)
                break

        if price_i is None:
            # fallback: pick the first column that contains "price"
            for i, h in enumerate(headers):
                if "price" in h:
                    price_i = i
                    break

        if price_i is None:
            return

        # Iterate data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            try:
                name = str(row[name_i]).strip() if row[name_i] is not None else ""
            except Exception:
                name = ""
            if len(name) < 3:
                continue

            price_val = row[price_i] if price_i < len(row) else None
            price = None
            if isinstance(price_val, (int, float)):
                price = float(price_val)
            else:
                # try parse "$12.34"
                try:
                    s = str(price_val).replace("$", "").replace(",", "").strip()
                    price = float(s)
                except Exception:
                    price = None

            n = _norm(name)
            if not n:
                continue

            # Keep max price if duplicates
            if price is not None:
                existing = self._prices.get(n)
                if existing is None or price > existing:
                    self._prices[n] = price
                    self._source_by_name[n] = source_name
            else:
                self._prices.setdefault(n, None)
                self._source_by_name.setdefault(n, source_name)


@lru_cache(maxsize=1)
def get_price_index() -> PriceIndex:
    xlsx_dir = os.environ.get("CARD_XLSX_DIR")
    if not xlsx_dir:
        xlsx_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "data", "xlsx")
        xlsx_dir = os.path.abspath(xlsx_dir)

    idx = PriceIndex(xlsx_dir=xlsx_dir)
    idx.load()
    return idx


def lookup_price(card_name: str) -> Optional[PriceHit]:
    idx = get_price_index()
    idx.load()

    if not idx._names:
        return None

    q = _norm(card_name)
    if not q:
        return None

    # Exact
    if q in idx._prices:
        return PriceHit(name=card_name, price_raw=idx._prices.get(q), score=100, source=idx._source_by_name.get(q, "xlsx"))

    # Fuzzy
    match = process.extractOne(q, idx._names, scorer=fuzz.WRatio)
    if not match:
        return None

    best_name_norm, score, _ = match
    if score < 78:
        return None

    price = idx._prices.get(best_name_norm)
    source = idx._source_by_name.get(best_name_norm, "xlsx")

    return PriceHit(name=best_name_norm, price_raw=price, score=int(score), source=source)
